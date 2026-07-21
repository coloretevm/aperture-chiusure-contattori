"""Excel report generation with openpyxl."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from models import AnalysisConfig, AnalysisResult
from utils import format_date, format_duration, make_report_base_name, median

TITLE_FILL = PatternFill("solid", fgColor="17365D")
SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
ALT_FILL = PatternFill("solid", fgColor="F3F7FB")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


def create_excel_report(result: AnalysisResult, output_dir: Path) -> Path:
    """Create the professional Excel report and return its path."""

    if result.start_time is None or result.end_time is None:
        raise ValueError("Nessun periodo dati disponibile per il report.")

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{make_report_base_name(result.rtu_name, result.start_time, result.end_time)}.xlsx"

    workbook = Workbook()
    summary = workbook.active
    summary.title = "RIEPILOGO"
    _write_summary(summary, result)
    _write_events(workbook.create_sheet("APERTURE_E_CHIUSURE"), result)
    _write_isolated(workbook.create_sheet("INCREMENTI_ISOLATI"), result)
    _write_logic(workbook.create_sheet("LOGICA_USATA"), result.config)
    _write_original_data(workbook.create_sheet("DATI_ORIGINALI"), result)
    workbook.save(path)
    return path


def _write_summary(sheet, result: AnalysisResult) -> None:
    sheet["A1"] = f"ANALISI APERTURE E CHIUSURE - RTU {result.rtu_name}"
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    sheet.merge_cells("A1:D1")
    sheet["A2"] = (
        f"APERTURE E CHIUSURE RILEVATE DAL {format_date(result.start_time)} "
        f"AL {format_date(result.end_time)}"
    )
    sheet["A2"].fill = SUBTITLE_FILL
    sheet.merge_cells("A2:D2")

    intervals = [reading.interval_minutes for reading in result.readings if reading.interval_minutes is not None]
    counts = {
        "ALTA": sum(1 for event in result.events if event.reliability == "ALTA"),
        "MEDIA": sum(1 for event in result.events if event.reliability == "MEDIA"),
        "DA VERIFICARE": sum(1 for event in result.events if event.reliability == "DA VERIFICARE"),
    }
    rows = [
        ("Nome file analizzato", result.source_file.name),
        ("Nome GDC/RTU", result.rtu_name),
        ("Data iniziale dati", result.start_time),
        ("Data finale dati", result.end_time),
        ("Messaggi analizzati", len(result.readings)),
        ("Righe scartate", result.discarded_rows),
        ("Intervallo mediano tra messaggi (min)", median([float(item) for item in intervals]) or 0),
        ("Aperture rilevate", len(result.events)),
        ("Incrementi isolati", len(result.isolated_increments)),
        ("Consumo totale (m3)", round(result.total_consumption, 3)),
        ("Aperture con affidabilita alta", counts["ALTA"]),
        ("Aperture con affidabilita media", counts["MEDIA"]),
        ("Aperture da verificare", counts["DA VERIFICARE"]),
        ("Reset o anomalie rilevate", result.reset_count),
    ]
    for row_index, (label, value) in enumerate(rows, start=4):
        sheet.cell(row_index, 1, label).font = BOLD_FONT
        cell = sheet.cell(row_index, 2, value)
        if hasattr(value, "year"):
            cell.number_format = "dd/mm/yyyy hh:mm:ss.000"

    _autosize(sheet)


def _write_events(sheet, result: AnalysisResult) -> None:
    sheet["A1"] = f"RTU: {result.rtu_name}"
    sheet["A1"].font = Font(bold=True, size=13, color="17365D")
    sheet["A2"] = (
        f"APERTURE E CHIUSURE RILEVATE DAL {format_date(result.start_time)} "
        f"AL {format_date(result.end_time)}"
    )
    sheet["A2"].fill = SUBTITLE_FILL
    sheet["A3"] = (
        "La chiusura viene registrata sull'ultima variazione del contatore; "
        "i 90 minuti successivi servono esclusivamente a confermarla."
    )
    headers = [
        "N.",
        "DATA E ORA APERTURA",
        "Contatore iniziale (m3)",
        "DATA E ORA CHIUSURA",
        "Contatore finale (m3)",
        "CONSUMO (m3)",
        "SEPARATORE",
        "DURATA APERTURA",
        "Tipo apertura",
        "Affidabilita",
        "Incrementi positivi",
        "Intervallo max dati (min)",
        "Note",
        "Riga apertura",
        "Riga chiusura",
    ]
    header_row = 5
    sheet.append([])
    sheet.append(headers)
    for event in result.events:
        sheet.append(
            [
                event.number,
                event.opening_time,
                event.initial_counter,
                event.closing_time,
                event.final_counter,
                event.consumption,
                "",
                format_duration(event.duration_seconds),
                event.opening_type,
                event.reliability,
                event.positive_increments,
                event.max_interval_minutes,
                event.notes,
                event.opening_row,
                event.closing_row,
            ]
        )

    _style_table(sheet, header_row, max(header_row, sheet.max_row), len(headers), "TabAperture")
    for row in range(header_row + 1, sheet.max_row + 1):
        sheet.cell(row, 2).number_format = "dd/mm/yyyy hh:mm:ss.000"
        sheet.cell(row, 4).number_format = "dd/mm/yyyy hh:mm:ss.000"
        for col in (3, 5, 6):
            sheet.cell(row, col).number_format = "0.0"
        reliability_cell = sheet.cell(row, 10)
        if reliability_cell.value == "ALTA":
            reliability_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif reliability_cell.value == "MEDIA":
            reliability_cell.fill = PatternFill("solid", fgColor="FFEB9C")
        elif reliability_cell.value == "DA VERIFICARE":
            reliability_cell.fill = PatternFill("solid", fgColor="FFC7CE")

    if sheet.max_row > header_row:
        sheet.conditional_formatting.add(
            f"F{header_row + 1}:F{sheet.max_row}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color="5B9BD5"),
        )
    sheet.freeze_panes = "A6"
    _autosize(sheet)
    sheet.column_dimensions["G"].width = 4
    sheet.cell(header_row, 7).font = Font(color="1F4E78", bold=True)


def _write_isolated(sheet, result: AnalysisResult) -> None:
    headers = [
        "N.",
        "Data e ora iniziale",
        "Contatore iniziale",
        "Data e ora finale",
        "Contatore finale",
        "Incremento",
        "Passi positivi",
        "Motivo",
        "Righe originali",
    ]
    sheet.append(headers)
    for item in result.isolated_increments:
        sheet.append(
            [
                item.number,
                item.start_time,
                item.start_counter,
                item.end_time,
                item.end_counter,
                item.increment,
                item.positive_steps,
                item.reason,
                item.original_rows,
            ]
        )
    _style_table(sheet, 1, max(1, sheet.max_row), len(headers), "TabIncrementi")
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 2).number_format = "dd/mm/yyyy hh:mm:ss.000"
        sheet.cell(row, 4).number_format = "dd/mm/yyyy hh:mm:ss.000"
        for col in (3, 5, 6):
            sheet.cell(row, col).number_format = "0.0"
    sheet.freeze_panes = "A2"
    _autosize(sheet)


def _write_logic(sheet, config: AnalysisConfig) -> None:
    rows = [
        ("Correzione oscillazioni", f"Diminuzioni fino a {config.max_oscillation} m3 vengono mantenute al valore corretto precedente."),
        ("Apertura normale", f"Almeno {config.normal_opening_volume} m3 entro {config.normal_opening_window_minutes} minuti."),
        ("Apertura lenta", f"Almeno {config.slow_opening_volume} m3 entro {config.slow_opening_window_minutes} minuti con almeno {config.slow_opening_min_increments} incrementi positivi."),
        ("Apertura rapida", f"Un singolo incremento pari o superiore a {config.rapid_single_increment} m3."),
        ("Chiusura", f"Confermata dopo {config.closure_confirmation_minutes} minuti senza nuova sequenza di apertura."),
        ("Tolleranza chiusura", f"Consumo massimo tollerato durante la conferma: {config.closure_tolerance} m3."),
        ("Incrementi isolati", "Incrementi piccoli senza continuita sufficiente non vengono conteggiati come aperture."),
        ("Reset", "Una diminuzione superiore alla soglia di oscillazione apre un nuovo segmento di analisi."),
        ("Affidabilita", f"ALTA se chiusura confermata, nessuna anomalia e intervalli dati non superiori a {config.high_reliability_max_interval_minutes} minuti."),
    ]
    sheet.append(["Parametro", "Logica applicata"])
    for row in rows:
        sheet.append(list(row))
    _style_table(sheet, 1, sheet.max_row, 2, "TabLogica")
    _autosize(sheet)


def _write_original_data(sheet, result: AnalysisResult) -> None:
    headers = [
        "Riga originale",
        "Data e ora",
        "V1Counter originale",
        "V1Counter corretto",
        "Delta",
        "Intervallo minuti",
        "Anomalia",
        "Segmento",
    ]
    sheet.append(headers)
    for reading in result.readings:
        sheet.append(
            [
                reading.original_row,
                reading.timestamp,
                reading.original_value,
                reading.corrected_value,
                reading.delta,
                reading.interval_minutes,
                reading.anomaly,
                reading.segment,
            ]
        )
    _style_table(sheet, 1, sheet.max_row, len(headers), "TabOriginali")
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 2).number_format = "dd/mm/yyyy hh:mm:ss.000"
        for col in (3, 4, 5):
            sheet.cell(row, col).number_format = "0.0"
    sheet.freeze_panes = "A2"
    _autosize(sheet)


def _style_table(sheet, header_row: int, last_row: int, last_col: int, name: str) -> None:
    for col in range(1, last_col + 1):
        cell = sheet.cell(header_row, col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(header_row + 1, last_row + 1):
        if row % 2 == 0:
            for col in range(1, last_col + 1):
                sheet.cell(row, col).fill = ALT_FILL
        for col in range(1, last_col + 1):
            sheet.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
    if last_row >= header_row:
        reference = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
        table = Table(displayName=name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
        sheet.add_table(table)


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 45)
