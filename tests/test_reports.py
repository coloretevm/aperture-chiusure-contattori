from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

from analyzer import analyze_readings
from app import APP_ICON_NAME, APP_ICON_SOURCE_NAME, default_output_directory, enable_high_dpi_awareness, resource_path
from excel_report import create_excel_report
from models import AnalysisConfig, CounterReading
from openpyxl import load_workbook
from PIL import Image
from pdf_report import create_pdf_report


def test_excel_and_pdf_reports_are_created(tmp_path) -> None:
    start = datetime(2026, 6, 10, 17, 7, 57, 636000)
    end = datetime(2026, 6, 10, 20, 32, 45, 584000)
    readings = [
        CounterReading(2, start, 0.3),
        CounterReading(3, end, 3.2),
        CounterReading(4, end + timedelta(minutes=95), 3.2),
    ]
    result = analyze_readings(
        readings,
        AnalysisConfig(),
        Path("known.csv"),
        "CBG_0087",
        "CBG_0087 - V1Counter",
    )

    excel_path = create_excel_report(result, tmp_path)
    pdf_path = create_pdf_report(result, tmp_path)

    assert excel_path.exists()
    assert pdf_path.exists()
    assert excel_path.suffix == ".xlsx"
    assert pdf_path.suffix == ".pdf"


def test_default_output_directory_prefers_desktop() -> None:
    desktop = Path.home() / "Desktop"
    expected = desktop if desktop.exists() else Path.home()

    output_dir = default_output_directory()

    assert output_dir == expected


def test_high_dpi_awareness_setup_is_callable() -> None:
    enable_high_dpi_awareness()


def test_app_icon_resource_exists() -> None:
    assert resource_path(APP_ICON_NAME).exists()


def test_app_icon_png_source_exists() -> None:
    source = resource_path(APP_ICON_SOURCE_NAME)
    image = Image.open(source)

    assert source.exists()
    assert image.size == (256, 256)
    assert image.mode == "RGBA"


def test_app_icon_contains_taskbar_sizes() -> None:
    icon = Image.open(resource_path(APP_ICON_NAME))

    assert {(256, 256), (128, 128), (64, 64), (48, 48), (32, 32), (24, 24), (16, 16)}.issubset(icon.ico.sizes())


def test_excel_event_table_has_no_empty_headers(tmp_path) -> None:
    start = datetime(2026, 6, 10, 17, 7, 57, 636000)
    end = datetime(2026, 6, 10, 20, 32, 45, 584000)
    readings = [
        CounterReading(2, start, 0.3),
        CounterReading(3, end, 3.2),
        CounterReading(4, end + timedelta(minutes=95), 3.2),
    ]
    result = analyze_readings(
        readings,
        AnalysisConfig(),
        Path("known.csv"),
        "CBG_0087",
        "CBG_0087 - V1Counter",
    )

    excel_path = create_excel_report(result, tmp_path)
    workbook = load_workbook(excel_path)
    sheet = workbook["APERTURE_E_CHIUSURE"]

    assert all(sheet.cell(5, column).value is not None for column in range(1, 16))
